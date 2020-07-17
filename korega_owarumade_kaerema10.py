import openpyxl
from tkinter import messagebox
import tkinter.simpledialog as simpledialog
import tkinter as tk

root = tk.Tk()
root.withdraw()#小さなウィンドウを表示させない
root.title(u"終わるまで帰れま10")
root.geometry("600x600")
#root.mainloop()

print('プログラムを起動中...')
#エクセルからメッセージを読み込んでメッセージリスト作成
message_list = []
wb = openpyxl.load_workbook('message_box.xlsx')
sheet = wb['Sheet1']
for i in range(2, sheet.max_row):
    message = sheet.cell(row=i, column=1).value
    message_list.append(message)

#メッセージボックス
for i in message_list:
    while True:
        ret = messagebox.askyesno('確認', i)
        if ret == False:
            messagebox.showerror('警告', '＜＜今すぐにタスクを終わらせましょう＞＞\n' + i)
        else:
            break
inputdata = simpledialog.askstring("引き継ぎ事項", "引き継ぎ事項はありますか？",)
if inputdata:
    with open('hikitugi.txt', 'w', encoding='utf-8') as f:
        f.write(inputdata)
    messagebox.showinfo('引き継ぎ完了', 'hikitugi.txtに引き継ぎ事項を書き込みました！')
messagebox.showinfo('おめでとう！', 'お疲れ様です。早く帰りましょう！！')
